"""アンケート結果を Excel ファイルとして出力する。

シート構成:
  - 質問別: 質問ごとに [P.N] 付きで回答を列挙
  - 回答者別マトリクス: 行=回答者、列=質問の表
  - 定量サマリー: センチメント・トピック・キーワード（与えられた場合）
"""
from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from survey_analyzer import ParsedQuestion, RespondentBlock


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(cell) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_width: int = 12, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row:
                if c.value is None:
                    continue
                text = str(c.value)
                for line in text.splitlines():
                    longest = max(longest, min(max_width, int(len(line) * 1.6)))
        ws.column_dimensions[letter].width = longest


def build_workbook(
    questions: list[ParsedQuestion],
    respondents: list[RespondentBlock],
    quant_summary: Optional[dict] = None,
    source_filename: str = "",
) -> bytes:
    """質問別・回答者別マトリクス・定量サマリーを含むワークブックを生成."""
    wb = Workbook()

    # 1) 質問別シート
    ws_q = wb.active
    ws_q.title = "質問別"
    ws_q.append(["質問番号", "質問文", "回答者ID", "回答"])
    for c in ws_q[1]:
        _style_header(c)
    for q in questions:
        if not q.answers:
            ws_q.append([f"Q{q.q_num}", q.q_text, "", "（無回答）"])
        for ans in q.answers:
            ws_q.append([f"Q{q.q_num}", q.q_text, ans.respondent_id, ans.text])
    for row in ws_q.iter_rows(min_row=2):
        for c in row:
            c.alignment = _WRAP
    _auto_width(ws_q)

    # 2) 回答者別マトリクス
    ws_r = wb.create_sheet("回答者別マトリクス")
    q_nums = [q.q_num for q in questions]
    q_text_lookup = {q.q_num: q.q_text for q in questions}
    header = ["回答者ID"] + [f"Q{n}: {q_text_lookup[n]}" for n in q_nums]
    ws_r.append(header)
    for c in ws_r[1]:
        _style_header(c)
    for block in respondents:
        ans_map = {q_num: a for q_num, _q_text, a in block.answers}
        row = [block.respondent_id]
        for n in q_nums:
            row.append(ans_map.get(n, ""))
        ws_r.append(row)
    for row in ws_r.iter_rows(min_row=2):
        for c in row:
            c.alignment = _WRAP
    _auto_width(ws_r)

    # 3) 定量サマリーシート
    if quant_summary:
        ws_s = wb.create_sheet("定量サマリー")
        _write_summary_sheet(ws_s, quant_summary, source_filename)
        _auto_width(ws_s)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_summary_sheet(ws, summary: dict, source_filename: str) -> None:
    if source_filename:
        ws.append(["元ファイル", source_filename])
        ws.append([])

    ws.append(["分析対象回答数", summary.get("answer_count", "")])
    ws.append([])

    ws.append(["センチメント分類"])
    _style_header(ws.cell(row=ws.max_row, column=1))
    sent = summary.get("sentiment", {}) or {}
    ws.append(["ポジティブ", sent.get("positive", 0)])
    ws.append(["ネガティブ", sent.get("negative", 0)])
    ws.append(["中立", sent.get("neutral", 0)])
    ws.append([])

    topics = summary.get("topics", []) or []
    if topics:
        ws.append(["トピック", "件数", "代表表現"])
        for c in ws[ws.max_row]:
            _style_header(c)
        for t in topics:
            examples = " / ".join(t.get("examples", []) or [])
            ws.append([t.get("label", ""), t.get("count", 0), examples])
        ws.append([])

    keywords = summary.get("keywords", []) or []
    if keywords:
        ws.append(["キーワード", "出現数"])
        for c in ws[ws.max_row]:
            _style_header(c)
        for k in keywords:
            ws.append([k.get("word", ""), k.get("count", 0)])
        ws.append([])

    highlights = summary.get("highlights", {}) or {}
    pos = highlights.get("positive", []) or []
    neg = highlights.get("negative", []) or []
    if pos or neg:
        ws.append(["代表的な回答"])
        _style_header(ws.cell(row=ws.max_row, column=1))
        for s in pos:
            ws.append(["ポジティブ", s])
        for s in neg:
            ws.append(["ネガティブ", s])

    for row in ws.iter_rows():
        for c in row:
            if c.alignment is None or not c.alignment.wrap_text:
                c.alignment = _WRAP
